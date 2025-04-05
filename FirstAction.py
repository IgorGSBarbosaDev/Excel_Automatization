from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelProcessor:
    @staticmethod
    def process_excel(file_path, output_path):
        """
        Processa o arquivo Excel conforme as instruções:
        1. Insere uma nova coluna no índice D e preenche com a fórmula =INT(C2).
        2. Copia e cola os valores da coluna D.
        3. Insere uma nova coluna no índice AB e preenche com a fórmula =CONCAT(AB;" ";AC).
        4. Copia e cola os valores da coluna AB.
        
        :param file_path: Caminho do arquivo Excel de entrada (.xls ou .xlsx).
        :param output_path: Caminho do arquivo Excel de saída (.xlsx).
        """
        # Carregar o arquivo Excel
        wb = load_workbook(file_path)
        ws = wb.active

        # Passo 1: Inserir uma nova coluna no índice D
        ws.insert_cols(4)  # Índice D é a coluna 4
        ws["D1"] = ""  # Deixar a célula D1 vazia

        # Preencher a coluna D com a fórmula =INT(C2)
        for row in range(2, ws.max_row + 1):
            cell_c = ws[f"C{row}"].value
            if isinstance(cell_c, (int, float)):  # Apenas números
                ws[f"D{row}"] = f"=INT(C{row})"

        # Copiar e colar os valores da coluna D
        for row in range(2, ws.max_row + 1):
            ws[f"D{row}"].value = ws[f"D{row}"].value

        # Passo 2: Inserir uma nova coluna no índice AB
        col_ab_index = 28  # Índice AB é a coluna 28
        ws.insert_cols(col_ab_index)
        ws[f"{get_column_letter(col_ab_index)}1"] = ""  # Deixar a célula AB1 vazia

        # Preencher a coluna AB com a fórmula =CONCAT(AB;" ";AC)
        for row in range(2, ws.max_row + 1):
            cell_ab = ws[f"AB{row}"].value
            cell_ac = ws[f"AC{row}"].value
            if cell_ab and cell_ac:  # Apenas se ambas as células tiverem valores
                ws[f"{get_column_letter(col_ab_index)}{row}"] = f'=CONCAT(AB{row};" ";AC{row})'

        # Copiar e colar os valores da coluna AB
        for row in range(2, ws.max_row + 1):
            ws[f"{get_column_letter(col_ab_index)}{row}"].value = ws[f"{get_column_letter(col_ab_index)}{row}"].value

        # Salvar o arquivo processado
        wb.save(output_path)
        print(f"Arquivo processado e salvo em: {output_path}")


# Exemplo de uso
if __name__ == "__main__":
    input_path = "caminho_para_seu_arquivo.xls"  # Substitua pelo caminho do arquivo de entrada
    output_path = "caminho_para_seu_arquivo_modificado.xlsx"  # Substitua pelo caminho do arquivo de saída
    ExcelProcessor.process_excel(input_path, output_path)