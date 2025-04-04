import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import win32com.client as win32

def convert_xls_to_xlsx(input_file, output_file):
    """
    Converte um arquivo .xls para .xlsx usando o Microsoft Excel via pywin32.
    """
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    try:
        # Abrir o arquivo .xls
        wb = excel.Workbooks.Open(input_file)

        # Salvar como .xlsx
        wb.SaveAs(output_file, FileFormat=51)  # 51 é o formato para .xlsx
        wb.Close()
    except Exception as e:
        raise ValueError(f"Erro ao converter o arquivo .xls para .xlsx: {e}")
    finally:
        excel.Application.Quit()

class ExcelProcessor:
    @staticmethod
    def process_excel(file_path, output_path):
        """
        Processa o arquivo Excel, adicionando e preenchendo colunas específicas.
        """
        # Verificar a extensão do arquivo de entrada
        file_extension = os.path.splitext(file_path)[1].lower()

        # Converter .xls para .xlsx se necessário
        if file_extension == '.xls':
            temp_output = file_path.replace('.xls', '.xlsx')
            convert_xls_to_xlsx(file_path, temp_output)
            file_path = temp_output
            file_extension = '.xlsx'

        # Escolher o mecanismo apropriado
        if file_extension == '.xlsx':
            df = pd.read_excel(file_path, engine='openpyxl')
        else:
            raise ValueError("Formato de arquivo não suportado. Use arquivos .xls ou .xlsx.")

        # Remover espaços extras dos nomes das colunas
        df.columns = df.columns.str.strip()

        # Verificar se a coluna 'Assunto Matrícula' existe
        if 'Assunto Matrícula' not in df.columns:
            raise ValueError("A coluna 'Assunto Matrícula' não foi encontrada no arquivo Excel.")

        # Inserir nova coluna na posição D (índice 3)
        df.insert(3, 'Nova Coluna D', None)

        # Preencher a coluna D com a função INT aplicada à coluna 'Assunto Matrícula'
        df['Nova Coluna D'] = df['Assunto Matrícula'].apply(
            lambda x: int(str(x).replace("T_", "")) if pd.notnull(x) and str(x).replace("T_", "").isdigit() else None
        )

        # Substituir a coluna D pelos valores calculados (remover fórmulas)
        df['Nova Coluna D'] = df['Nova Coluna D'].astype(object)

        # Inserir nova coluna na posição AB (índice 27)
        df.insert(27, 'Nova Coluna AB', None)

        # Preencher a coluna AB com a concatenação das colunas 'Nome do proprietário atual' e 'Sobrenome do proprietário atual'
        if 'Nome do proprietário atual' in df.columns and 'Sobrenome do proprietário atual' in df.columns:
            df['Nova Coluna AB'] = df.apply(
                lambda row: f"{row['Nome do proprietário atual']} {row['Sobrenome do proprietário atual']}" 
                if pd.notnull(row['Nome do proprietário atual']) and pd.notnull(row['Sobrenome do proprietário atual']) else None,
                axis=1
            )
        else:
            raise ValueError("As colunas 'Nome do proprietário atual' e 'Sobrenome do proprietário atual' não foram encontradas no arquivo Excel.")

        # Substituir a coluna AB pelos valores calculados (remover fórmulas)
        df['Nova Coluna AB'] = df['Nova Coluna AB'].astype(object)

        # Salvar a planilha modificada
        df.to_excel(output_path, index=False, engine='openpyxl')

    @staticmethod
    def apply_procv(target_file, processed_file):
        """
        Aplica a fórmula PROCV na coluna C da planilha de destino.
        """
        # Abrir a planilha de destino
        wb = openpyxl.load_workbook(target_file)
        ws = wb.active

        # Nome da aba da planilha processada
        processed_sheet_name = "Report"

        # Aplicar a fórmula PROCV na coluna C
        for row in range(2, ws.max_row + 1):  # Começa na linha 2
            cell = ws[f"C{row}"]
            cell.value = f'=PROCV(D{row};\'[{processed_file}]' + f'{processed_sheet_name}\'!$D:$M;10;0)'

        # Salvar a planilha de destino
        wb.save(target_file)

class GUI:
    def __init__(self, root):
        """
        Inicializa a interface gráfica.
        """
        self.root = root
        self.root.title("Processador de Excel")

        # Criar widgets
        self.label_file_path = tk.Label(root, text="Arquivo de entrada:")
        self.label_file_path.pack(pady=10)

        self.entry_file_path = tk.Entry(root, width=50)
        self.entry_file_path.pack(pady=5)

        self.button_browse = tk.Button(root, text="Procurar", command=self.select_file)
        self.button_browse.pack(pady=5)

        self.label_target_path = tk.Label(root, text="Arquivo de destino:")
        self.label_target_path.pack(pady=10)

        self.entry_target_path = tk.Entry(root, width=50)
        self.entry_target_path.pack(pady=5)

        self.button_browse_target = tk.Button(root, text="Procurar", command=self.select_target_file)
        self.button_browse_target.pack(pady=5)

        self.button_process = tk.Button(root, text="Executar Processo", command=self.run_process)
        self.button_process.pack(pady=20)

    def select_file(self):
        """
        Abre o diálogo para selecionar o arquivo de entrada.
        """
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            self.entry_file_path.delete(0, tk.END)
            self.entry_file_path.insert(0, file_path)

    def select_target_file(self):
        """
        Abre o diálogo para selecionar o arquivo de destino.
        """
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.entry_target_path.delete(0, tk.END)
            self.entry_target_path.insert(0, file_path)

    def run_process(self):
        """
        Executa o processamento do arquivo Excel.
        """
        input_file = self.entry_file_path.get()
        target_file = self.entry_target_path.get()
        output_file = input_file.replace('.xls', '_modificado.xlsx').replace('.xlsx', '_modificado.xlsx')

        try:
            # Processar a planilha original
            ExcelProcessor.process_excel(input_file, output_file)

            # Aplicar a fórmula PROCV na planilha de destino
            ExcelProcessor.apply_procv(target_file, output_file)

            messagebox.showinfo("Sucesso", f"Processo concluído! Arquivo salvo como: {output_file}")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = GUI(root)
    root.mainloop()