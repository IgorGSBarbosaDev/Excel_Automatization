from openpyxl import load_workbook
from datetime import datetime

class ExcelUpdater:
    @staticmethod
    def update_official_sheet(converted_file, official_file, sheet_name="Report"):
        """
        Atualiza as colunas "Etapa", "Proprietário Atual", "Nota atribuída" e "Potencial Atribuído" na PlanilhaOficial com fórmulas PROCV.
        Salva o arquivo atualizado com o padrão de nome: dataatual(ano, mes, dia)_Status_Ciclo_2024_BC.xlsx.
        
        :param converted_file: Caminho do arquivo convertido (.xlsx).
        :param official_file: Caminho da PlanilhaOficial (.xlsx).
        :param sheet_name: Nome da aba no arquivo convertido que contém os dados (padrão: "Report").
        """
        # Abrir a PlanilhaOficial
        wb_official = load_workbook(official_file)
        ws_official = wb_official.active

        # Função auxiliar para encontrar a coluna pelo cabeçalho
        def find_column_by_header(header_name):
            for col in ws_official.iter_cols(1, ws_official.max_column):
                if col[0].value == header_name:
                    return col[0].column
            raise ValueError(f"A coluna '{header_name}' não foi encontrada na PlanilhaOficial.")

        # Atualizar a coluna "Etapa"
        etapa_col = find_column_by_header("Etapa")
        for row in range(2, ws_official.max_row + 1):
            ws_official.cell(row=row, column=etapa_col).value = None  # Limpar valores
            ws_official.cell(row=row, column=etapa_col).value = (
                f'=PROCV(D{row};\'[{converted_file.split("\\")[-1]}]{sheet_name}\'!$D:$M;10;0)'
            )

        # Atualizar a coluna "Proprietário Atual"
        proprietario_col = find_column_by_header("Proprietário Atual")
        for row in range(2, ws_official.max_row + 1):
            ws_official.cell(row=row, column=proprietario_col).value = None  # Limpar valores
            ws_official.cell(row=row, column=proprietario_col).value = (
                f'=PROCV(D{row};\'[{converted_file.split("\\")[-1]}]{sheet_name}\'!$D:$AB;25;0)'
            )

        # Atualizar a coluna "Nota atribuída"
        nota_col = find_column_by_header("Nota atribuída")
        for row in range(2, ws_official.max_row + 1):
            ws_official.cell(row=row, column=nota_col).value = None  # Limpar valores
            ws_official.cell(row=row, column=nota_col).value = (
                f'=PROCV(D{row};\'[{converted_file.split("\\")[-1]}]{sheet_name}\'!$D:$Z;23;0)'
            )

        # Atualizar a coluna "Potencial Atribuído"
        potencial_col = find_column_by_header("Potencial Atribuído")
        for row in range(2, ws_official.max_row + 1):
            ws_official.cell(row=row, column=potencial_col).value = None  # Limpar valores
            ws_official.cell(row=row, column=potencial_col).value = (
                f'=PROCV(D{row};\'[{converted_file.split("\\")[-1]}]{sheet_name}\'!$D:$Y;22;0)'
            )

        # Gerar o nome do arquivo com base na data atual
        current_date = datetime.now().strftime("%Y%m%d")
        new_file_name = f"{current_date}_Status_Ciclo_2024_BC.xlsx"

        # Salvar o arquivo atualizado com o novo nome
        wb_official.save(new_file_name)
        print(f"PlanilhaOficial atualizada e salva como: {new_file_name}")


# Exemplo de uso
if __name__ == "__main__":
    converted_path = "caminho_para_arquivo_convertido.xlsx"  # Substitua pelo caminho do arquivo convertido
    official_path = "caminho_para_planilha_oficial.xlsx"  # Substitua pelo caminho da PlanilhaOficial
    ExcelUpdater.update_official_sheet(converted_path, official_path)